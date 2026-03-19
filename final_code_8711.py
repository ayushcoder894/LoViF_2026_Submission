"""
LoViF 2026 - Semantic Quality Assessment
Solution v23: Optuna Hyperparameter Tuned using pyiqa (TOPIQ, MUSIQ, CLIPIQA, PieAPP) + Caching
Target: >0.95 without overfitting

Upgrades over v6:
1. Extra Data Initialization: Adds SOTA metrics pre-trained on external IQA datasets (KADID, KonIQ, PIPAL).
2. Feature caching added to drastically speed up re-training.
3. Extra Data flag set to 1 in the final readme.
"""

import argparse
import json
import os
import pickle
import platform
import random
import time
import warnings
import zipfile
import numpy as np
from PIL import Image
from scipy.stats import spearmanr, pearsonr, skew, kurtosis
from skimage.metrics import structural_similarity as ssim

import torch
import torch.nn.functional as F

warnings.filterwarnings('ignore')

DEFAULT_SEED = 42

DEVICE = 'cuda' if torch.cuda.is_available() else 'cpu'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, '.feature_cache_v41_creative')
os.makedirs(CACHE_DIR, exist_ok=True)


def configure_reproducibility(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    try:
        torch.use_deterministic_algorithms(True, warn_only=True)
    except Exception:
        pass
    if hasattr(torch.backends, 'cudnn'):
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def ensure_exists(path, kind='path'):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Missing required {kind}: {path}")


def canonical_readme_bytes(extra_data_flag=1):
    text = (
        "runtime per video [s] : 5.0\r\n"
        "CPU[1] / GPU[0] : 0\r\n"
        f"Extra Data [1] / No Extra Data [0] : {extra_data_flag}\r\n"
        "Other description : FINAL SUBMISSION.\r\n"
    )
    return text.encode('utf-8')


def resolve_readme_bytes(readme_template_zip, extra_data_flag=1):
    if readme_template_zip and os.path.exists(readme_template_zip):
        with zipfile.ZipFile(readme_template_zip, 'r') as zf:
            return zf.read('readme.txt')
    return canonical_readme_bytes(extra_data_flag=extra_data_flag)


def write_submission_zip(zip_path, xlsx_path, readme_bytes):
    fixed_time = (2026, 1, 1, 0, 0, 0)
    with open(xlsx_path, 'rb') as f:
        xlsx_bytes = f.read()
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        xlsx_info = zipfile.ZipInfo('prediction.xlsx', date_time=fixed_time)
        xlsx_info.compress_type = zipfile.ZIP_DEFLATED
        zf.writestr(xlsx_info, xlsx_bytes)

        readme_info = zipfile.ZipInfo('readme.txt', date_time=fixed_time)
        readme_info.compress_type = zipfile.ZIP_DEFLATED
        zf.writestr(readme_info, readme_bytes)


def sha256_of_file(path):
    import hashlib

    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def validate_submission_artifacts(zip_path, expected_rows):
    import openpyxl

    with zipfile.ZipFile(zip_path, 'r') as zf:
        names = sorted(zf.namelist())
        expected_names = ['prediction.xlsx', 'readme.txt']
        if names != expected_names:
            raise RuntimeError(f"Zip content mismatch. Expected {expected_names}, got {names}")

        with zf.open('prediction.xlsx') as fp:
            wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        header = [ws['A1'].value, ws['B1'].value]
        if header != ['picture_name', 'Score']:
            raise RuntimeError(f"Invalid header row in prediction.xlsx: {header}")
        row_count = ws.max_row - 1
        if row_count != expected_rows:
            raise RuntimeError(f"Invalid row count in prediction.xlsx: expected {expected_rows}, got {row_count}")


def write_repro_manifest(manifest_path, args, device, train_count, test_count, metrics, xlsx_path, zip_path, elapsed_sec):
    payload = {
        'timestamp_utc': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
        'platform': platform.platform(),
        'python_version': platform.python_version(),
        'numpy_version': np.__version__,
        'torch_version': torch.__version__,
        'device': device,
        'seed': args.seed,
        'train_root': os.path.abspath(args.train_root),
        'test_root': os.path.abspath(args.test_root),
        'cache_dir': os.path.abspath(args.cache_dir),
        'output_dir': os.path.abspath(args.output_dir),
        'output_zip': os.path.abspath(args.output_zip),
        'readme_template_zip': args.readme_template_zip,
        'train_samples': int(train_count),
        'test_samples': int(test_count),
        'cv_srocc': float(metrics['srocc']),
        'cv_plcc': float(metrics['plcc']),
        'cv_target': float(metrics['target']),
        'prediction_xlsx_sha256': sha256_of_file(xlsx_path),
        'submission_zip_sha256': sha256_of_file(zip_path),
        'elapsed_seconds': float(elapsed_sec),
    }
    with open(manifest_path, 'w', encoding='utf-8', newline='\n') as f:
        json.dump(payload, f, indent=2)


def build_arg_parser():
    parser = argparse.ArgumentParser(description='Reproducible LoViF 2026 SIQA submission runner')
    parser.add_argument('--train-root', default=os.path.join(BASE_DIR, 'Train (1)', 'Train'))
    parser.add_argument('--test-root', default=os.path.join(BASE_DIR, 'test-input', 'Test'))
    parser.add_argument('--cache-dir', default=os.path.join(BASE_DIR, '.feature_cache_v41_creative'))
    parser.add_argument('--output-dir', default=os.path.join(BASE_DIR, 'v38_impr_submission'))
    parser.add_argument('--output-zip', default=os.path.join(BASE_DIR, 'prediction_v41_creative.zip'))
    parser.add_argument('--readme-template-zip', default=os.path.join(BASE_DIR, 'prediction_raw.zip'))
    parser.add_argument('--seed', type=int, default=DEFAULT_SEED)
    parser.add_argument('--device', choices=['auto', 'cpu', 'cuda'], default='auto')
    return parser

# =========================================================================
# Deep Models Setup
# =========================================================================

def load_models():
    print("Loading models...")
    models = {}
    
    # OpenCLIP Big Models
    import open_clip
    print("  -> CLIP ViT-L-14")
    clip_h, _, clip_h_pre = open_clip.create_model_and_transforms('ViT-L-14', pretrained='openai')
    models['clipH'] = clip_h.to(DEVICE).eval()
    models['clipH_pre'] = clip_h_pre
    
    # NEW: Text-Anchored Quality Prompts
    tokenizer = open_clip.get_tokenizer('ViT-L-14')
    text_good = tokenizer(["A pristine, perfectly sharp, high-resolution masterpiece photograph."]).to(DEVICE)
    text_bad = tokenizer(["A blurry, noisy, heavily compressed, distorted, low-quality image."]).to(DEVICE)
    with torch.no_grad():
        emb_good = models['clipH'].encode_text(text_good)
        emb_good /= emb_good.norm(dim=-1, keepdim=True)
        emb_bad = models['clipH'].encode_text(text_bad)
        emb_bad /= emb_bad.norm(dim=-1, keepdim=True)
    models['text_good'] = emb_good.cpu().numpy()[0]
    models['text_bad'] = emb_bad.cpu().numpy()[0]

    print("  -> CLIP ViT-B-16")
    clip_l, _, clip_l_pre = open_clip.create_model_and_transforms('ViT-B-16', pretrained='openai')
    models['clipL'] = clip_l.to(DEVICE).eval()
    models['clipL_pre'] = clip_l_pre

    # DINOv2
    print("  -> DINOv2 ViT-L-14")
    dino = torch.hub.load('facebookresearch/dinov2', 'dinov2_vitl14', trust_repo=True)
    models['dino'] = dino.to(DEVICE).eval()

    # Perceptual Models
    print("  -> LPIPS (Alex + VGG)")
    import lpips
    models['lpips_alex'] = lpips.LPIPS(net='alex').to(DEVICE).eval()
    models['lpips_vgg'] = lpips.LPIPS(net='vgg').to(DEVICE).eval()

    import importlib
    pyiqa = importlib.import_module('pyiqa')
    print("  -> PIQ (DISTS)")
    import piq
    models['piq_dists'] = piq.DISTS(reduction='none').to(DEVICE).eval()
    
    print("  -> EXTRA DATA MODELS (pyiqa): TOPIQ-FR, MUSIQ, CLIPIQA, PieAPP")
    models['topiq_fr'] = pyiqa.create_metric('topiq_fr', device=DEVICE).eval()
    models['musiq'] = pyiqa.create_metric('musiq', device=DEVICE).eval()
    models['clipiqa'] = pyiqa.create_metric('clipiqa', device=DEVICE).eval()
    models['pieapp'] = pyiqa.create_metric('pieapp', device=DEVICE).eval()
    
    return models

# =========================================================================
# Feature Extraction
# =========================================================================

@torch.no_grad()
def get_clip_emb(model, transform, pil_img):
    t = transform(pil_img).unsqueeze(0).to(DEVICE)
    emb = model.encode_image(t)
    emb = F.normalize(emb, p=2, dim=-1)
    return emb.cpu().numpy()[0]

@torch.no_grad()
def get_dino_features(model, pil_img, size=224):
    import torchvision.transforms as T
    transform = T.Compose([
        T.Resize((size, size), interpolation=T.InterpolationMode.LANCZOS),
        T.ToTensor(),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),
    ])
    t = transform(pil_img).unsqueeze(0).to(DEVICE)
    out = model.forward_features(t)
    
    cls_emb = F.normalize(out['x_norm_clstoken'], p=2, dim=-1)[0].cpu().numpy()
    
    # Patch features (1, 256, 1024)
    patches = out['x_norm_patchtokens'][0]
    patches = F.normalize(patches, p=2, dim=-1).cpu().numpy()
    
    return cls_emb, patches

def pil_to_torch_01(pil_img, size=(256, 256)):
    """Convert PIL to [1, C, H, W] in [0, 1] for LPIPS and PYIQA"""
    import torchvision.transforms.functional as TF
    t = TF.resize(pil_img, size, interpolation=TF.InterpolationMode.LANCZOS)
    t = TF.to_tensor(t).unsqueeze(0)
    return t.to(DEVICE)

def extract_dense_and_scalar_features(d_img, r_img, models):
    scalars = {}
    dense = {}

    import torchvision.transforms.functional as TF
    import piq

    # 1. Standardizations
    t_d_256 = pil_to_torch_01(d_img, (256, 256))
    t_r_256 = pil_to_torch_01(r_img, (256, 256))
    
    t_d_512 = pil_to_torch_01(d_img, (512, 512))
    t_r_512 = pil_to_torch_01(r_img, (512, 512))

    with torch.no_grad():
        # LPIPS
        lp_d, lp_r = t_d_256 * 2 - 1, t_r_256 * 2 - 1
        scalars['lpips_alex'] = models['lpips_alex'](lp_d, lp_r).item()
        scalars['lpips_vgg'] = models['lpips_vgg'](lp_d, lp_r).item()
        
        lp_d5, lp_r5 = t_d_512 * 2 - 1, t_r_512 * 2 - 1
        scalars['lpips_alex_512'] = models['lpips_alex'](lp_d5, lp_r5).item()
        scalars['lpips_vgg_512'] = models['lpips_vgg'](lp_d5, lp_r5).item()
        
        # PIQ Suite
        scalars['dists'] = models['piq_dists'](t_d_256, t_r_256).item()
        scalars['dists_512'] = models['piq_dists'](t_d_512, t_r_512).item()
        
        scalars['fsim'] = piq.fsim(t_d_256, t_r_256).item()
        scalars['gmsd'] = piq.gmsd(t_d_256, t_r_256).item()
        scalars['mdsi'] = piq.mdsi(t_d_256, t_r_256).item()
        scalars['haarpsi'] = piq.haarpsi(t_d_256, t_r_256).item()
        scalars['vsi'] = piq.vsi(t_d_256, t_r_256).item()
        scalars['srsim'] = piq.srsim(t_d_256, t_r_256).item()
        scalars['ms_ssim'] = piq.multi_scale_ssim(t_d_256, t_r_256).item()
        scalars['vif_p'] = piq.vif_p(t_d_256, t_r_256).item()
        
        # SOTA EXTRA DATA MODELS (Trained on external datasets)
        scalars['topiq_fr'] = models['topiq_fr'](t_d_256, t_r_256).item()
        scalars['musiq'] = models['musiq'](t_d_256).item()
        scalars['clipiqa'] = models['clipiqa'](t_d_256).item()
        scalars['pieapp'] = models['pieapp'](t_d_256, t_r_256).item()
        
    # 2. Extract CLIP H
    dh = get_clip_emb(models['clipH'], models['clipH_pre'], d_img)
    rh = get_clip_emb(models['clipH'], models['clipH_pre'], r_img)
    dense['clipH'] = np.concatenate([np.abs(dh - rh), dh * rh])
    scalars['clipH_cos'] = np.dot(dh, rh)
    scalars['clipH_l2'] = np.linalg.norm(dh - rh)
    scalars['clipH_l1'] = np.linalg.norm(dh - rh, ord=1)

    # NEW: Text-Anchored Structural Limits
    # Measures the specific cosine correlation to the conceptual idea of 'Good' vs 'Bad' limits
    scalars['clipH_text_good_d'] = np.dot(dh, models['text_good'])
    scalars['clipH_text_bad_d'] = np.dot(dh, models['text_bad'])
    scalars['clipH_text_good_r'] = np.dot(rh, models['text_good'])
    scalars['clipH_text_bad_r'] = np.dot(rh, models['text_bad'])
    # Relative perceptual drift vs idealized semantics
    scalars['clipH_text_diff_good'] = scalars['clipH_text_good_d'] - scalars['clipH_text_good_r']
    scalars['clipH_text_diff_bad'] = scalars['clipH_text_bad_d'] - scalars['clipH_text_bad_r']

    # 3. Extract CLIP L
    dl = get_clip_emb(models['clipL'], models['clipL_pre'], d_img)
    rl = get_clip_emb(models['clipL'], models['clipL_pre'], r_img)
    dense['clipL'] = np.concatenate([np.abs(dl - rl), dl * rl])
    scalars['clipL_cos'] = np.dot(dl, rl)
    scalars['clipL_l1'] = np.linalg.norm(dl - rl, ord=1)

    # 4. Extract DINOv2 (224)
    dd_cls, dd_patches = get_dino_features(models['dino'], d_img, size=224)
    rd_cls, rd_patches = get_dino_features(models['dino'], r_img, size=224)
    
    dense['dino'] = np.concatenate([np.abs(dd_cls - rd_cls), dd_cls * rd_cls])
    scalars['dino_cos'] = np.dot(dd_cls, rd_cls)
    scalars['dino_l2'] = np.linalg.norm(dd_cls - rd_cls)
    
    patch_cosines = (dd_patches * rd_patches).sum(axis=-1)
    dense['dino_patch_sim'] = patch_cosines
    scalars['dino_patch_mean'] = patch_cosines.mean()
    scalars['dino_patch_min'] = patch_cosines.min()
    scalars['dino_patch_std'] = patch_cosines.std()
    
    # NEW: Spatial Density Anomalies
    scalars['dino_patch_skew'] = float(np.nan_to_num(skew(patch_cosines)))
    scalars['dino_patch_kurt'] = float(np.nan_to_num(kurtosis(patch_cosines)))
    
    scalars['dino_patch_q10'] = np.percentile(patch_cosines, 10)
    scalars['dino_gap'] = scalars['dino_cos'] - scalars['dino_patch_min']

    # 5. Extract DINOv2 (336)
    dd_cls_336, dd_patches_336 = get_dino_features(models['dino'], d_img, size=336)
    rd_cls_336, rd_patches_336 = get_dino_features(models['dino'], r_img, size=336)
    dense['dino_336'] = np.concatenate([np.abs(dd_cls_336 - rd_cls_336), dd_cls_336 * rd_cls_336])

    patch_cosines_336 = (dd_patches_336 * rd_patches_336).sum(axis=-1)
    dense['dino_patch_sim_336'] = patch_cosines_336

    scalars['dino_cos_336'] = np.dot(dd_cls_336, rd_cls_336)
    scalars['dino_patch_mean_336'] = patch_cosines_336.mean()
    scalars['dino_patch_min_336'] = patch_cosines_336.min()
    
    # NEW: Spatial Entropy Metrics High-Res
    scalars['dino_patch_skew_336'] = float(np.nan_to_num(skew(patch_cosines_336)))
    scalars['dino_patch_kurt_336'] = float(np.nan_to_num(kurtosis(patch_cosines_336)))

    # 6. Traditional Pixel SSIM / MSE on Grayscale
    d_gray = np.array(d_img.convert('L').resize((256, 256)))
    r_gray = np.array(r_img.convert('L').resize((256, 256)))
    scalars['ssim_256'] = ssim(d_gray, r_gray, data_range=255)
    scalars['mse_256'] = np.mean((d_gray.astype(float) - r_gray.astype(float))**2)
    
    # NEW: Frequency Phase Error (FFT Mapping)
    f_d = np.fft.fft2(d_gray)
    f_r = np.fft.fft2(r_gray)
    scalars['fft_mag_mse'] = float(np.nan_to_num(np.mean((np.abs(f_d) - np.abs(f_r))**2)))
    scalars['fft_phase_mse'] = float(np.nan_to_num(np.mean((np.angle(f_d) - np.angle(f_r))**2)))
    return dense, scalars

def load_image_downscaled(path, max_dim=1024):
    img = Image.open(path).convert('RGB')
    if max(img.size) > max_dim:
        img.thumbnail((max_dim, max_dim), Image.BILINEAR)
    return img

def extract_features_cached(fname, d_path, r_path, models, flip=False):
    cache_path = os.path.join(CACHE_DIR, f"{fname}.pkl")
    if os.path.exists(cache_path):
        with open(cache_path, 'rb') as f:
            return pickle.load(f)
            
    d_img = load_image_downscaled(d_path)
    r_img = load_image_downscaled(r_path)
    
    if flip:
        d_img = d_img.transpose(Image.FLIP_LEFT_RIGHT)
        r_img = r_img.transpose(Image.FLIP_LEFT_RIGHT)
    
    dense, scalars = extract_dense_and_scalar_features(d_img, r_img, models)
    res = (dense, scalars)
    
    with open(cache_path, 'wb') as f:
        pickle.dump(res, f)
        
    return res

# =========================================================================
# Meta-Feature Generation
# =========================================================================

def train_dense_meta_models(X_dense_dict, y, folds):
    from sklearn.linear_model import Ridge
    oof_meta_features = {}
    fitted_meta_models = {}
    
    for model_key, X_mat in X_dense_dict.items():
        print(f"    -> Training Meta-Ridge on {model_key} diff vectors ({X_mat.shape[1]} dims)")
        oof_preds = np.zeros(len(y))
        
        for train_idx, val_idx in folds:
            X_tr, X_va = X_mat[train_idx], X_mat[val_idx]
            y_tr = y[train_idx]
            
            r = Ridge(alpha=500.0)
            r.fit(X_tr, y_tr)
            oof_preds[val_idx] = r.predict(X_va)
            
        oof_meta_features[model_key + '_ridge_pred'] = oof_preds
        
        final_r = Ridge(alpha=500.0)
        final_r.fit(X_mat, y)
        fitted_meta_models[model_key] = final_r
        
    return oof_meta_features, fitted_meta_models

def train_pairwise_rank_model(
    X_train,
    y_train,
    X_val=None,
    y_val=None,
    epochs=380,
    lambda_rank=0.80,
    lambda_mse=0.25,
    patience=35,
):
    """Train MLP with MSE + smooth pairwise ranking loss and optional early stopping."""
    y_mean = float(np.mean(y_train))
    y_std = float(np.std(y_train) + 1e-8)
    y_train_norm = (y_train - y_mean) / y_std

    x_t = torch.tensor(X_train, dtype=torch.float32, device=DEVICE)
    y_t = torch.tensor(y_train_norm, dtype=torch.float32, device=DEVICE)

    x_val_t = None
    y_val_np = None
    if X_val is not None and y_val is not None:
        x_val_t = torch.tensor(X_val, dtype=torch.float32, device=DEVICE)
        y_val_np = np.array(y_val, dtype=np.float32)

    model = torch.nn.Sequential(
        torch.nn.Linear(x_t.shape[1], 192),
        torch.nn.ReLU(),
        torch.nn.Dropout(0.12),
        torch.nn.Linear(192, 64),
        torch.nn.ReLU(),
        torch.nn.Linear(64, 1),
    ).to(DEVICE)

    optimizer = torch.optim.AdamW(model.parameters(), lr=2e-3, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=epochs)

    best_state = None
    best_score = -1e9
    bad_epochs = 0

    for _ in range(epochs):
        model.train()
        optimizer.zero_grad()

        pred = model(x_t).squeeze(1)
        mse_loss = F.mse_loss(pred, y_t)

        diff_pred = pred.unsqueeze(1) - pred.unsqueeze(0)
        diff_gt = y_t.unsqueeze(1) - y_t.unsqueeze(0)
        sign_gt = torch.sign(diff_gt)
        valid = sign_gt != 0
        # Smooth ranking loss is easier to optimize than hard hinge.
        rank_loss = F.softplus(-diff_pred * sign_gt)[valid].mean()

        loss = lambda_mse * mse_loss + lambda_rank * rank_loss
        loss.backward()
        optimizer.step()
        scheduler.step()

        if x_val_t is not None:
            model.eval()
            with torch.no_grad():
                val_pred_norm = model(x_val_t).squeeze(1).detach().cpu().numpy()
            val_pred = val_pred_norm * y_std + y_mean
            s = spearmanr(val_pred, y_val_np)[0]
            p = pearsonr(val_pred, y_val_np)[0]
            s = 0.0 if np.isnan(s) else float(s)
            p = 0.0 if np.isnan(p) else float(p)
            score = 0.75 * s + 0.25 * p

            if score > best_score:
                best_score = score
                best_state = {k: v.detach().clone() for k, v in model.state_dict().items()}
                bad_epochs = 0
            else:
                bad_epochs += 1
                if bad_epochs >= patience:
                    break

    if best_state is not None:
        model.load_state_dict(best_state)

    return {
        'model': model,
        'y_mean': y_mean,
        'y_std': y_std,
    }

@torch.no_grad()
def predict_pairwise_rank_model(model_pack, X_input):
    model = model_pack['model']
    y_mean = model_pack['y_mean']
    y_std = model_pack['y_std']
    x_t = torch.tensor(X_input, dtype=torch.float32, device=DEVICE)
    model.eval()
    pred_norm = model(x_t).squeeze(1).detach().cpu().numpy()
    return pred_norm * y_std + y_mean

# =========================================================================
# Main Execution
# =========================================================================

def main(args):
    global DEVICE, CACHE_DIR

    started = time.time()
    configure_reproducibility(args.seed)

    if args.device == 'auto':
        DEVICE = 'cuda' if torch.cuda.is_available() else 'cpu'
    else:
        if args.device == 'cuda' and not torch.cuda.is_available():
            raise RuntimeError('CUDA requested but not available on this machine.')
        DEVICE = args.device

    CACHE_DIR = os.path.abspath(args.cache_dir)
    os.makedirs(CACHE_DIR, exist_ok=True)

    train_root = os.path.abspath(args.train_root)
    test_root = os.path.abspath(args.test_root)

    train_dist_dir = os.path.join(train_root, 'Dist')
    train_ref_dir = os.path.join(train_root, 'Ref')
    scores_file = os.path.join(train_root, 'Train_scores.xlsx')

    target_dist_dir = os.path.join(test_root, 'Dist')
    target_ref_dir = os.path.join(test_root, 'Ref')

    ensure_exists(train_dist_dir, kind='train distorted directory')
    ensure_exists(train_ref_dir, kind='train reference directory')
    ensure_exists(scores_file, kind='training score file')
    ensure_exists(target_dist_dir, kind='test distorted directory')
    ensure_exists(target_ref_dir, kind='test reference directory')

    print("=" * 70)
    print("STEP 1: Load Ground Truth")
    import openpyxl
    wb = openpyxl.load_workbook(scores_file)
    ws = wb.active
    
    train_scores_map = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name, score = row[0], row[1]
        if name and score is not None:
            train_scores_map[str(name)] = float(score)
    print(f"Loaded {len(train_scores_map)} scores.")

    models = load_models()
    
    print("\n" + "=" * 70)
    print("STEP 2: Extract Training Features")
    train_files = sorted(os.listdir(train_dist_dir))
    train_files = [f for f in train_files if f in train_scores_map]
    
    X_scalars = []
    X_dense = {'clipH': [], 'clipL': [], 'dino': [], 'dino_patch_sim': [], 'dino_336': [], 'dino_patch_sim_336': []}
    y_train = []
    scalar_names = []
    
    for i, fname in enumerate(train_files):
        d_path = os.path.join(train_dist_dir, fname)
        r_path = os.path.join(train_ref_dir, fname)
        
        dense, scalars = extract_features_cached(f"train_{fname}", d_path, r_path, models, flip=False)
        
        if len(scalar_names) == 0:
            scalar_names = sorted(scalars.keys())
            
        X_scalars.append([scalars[k] for k in scalar_names])
        for k in X_dense.keys():
            X_dense[k].append(dense[k])
            
        y_train.append(train_scores_map[fname])
        
        if (i+1) % 50 == 0:
            print(f"  Processed {i+1} / {len(train_files)}")

    print("\n" + "=" * 70)
    print("STEP 3: Meta-Feature Engineering (Ridge Mappings)")

    X_scalars = np.nan_to_num(np.array(X_scalars))
    y_train = np.array(y_train)
    for k in X_dense:
        X_dense[k] = np.array(X_dense[k])
        
    from sklearn.model_selection import KFold
    folds = list(KFold(n_splits=5, shuffle=True, random_state=42).split(X_scalars))
    
    oof_meta, fitted_meta = train_dense_meta_models(X_dense, y_train, folds)
    
    for k, meta_array in oof_meta.items():
        X_scalars = np.hstack((X_scalars, meta_array.reshape(-1, 1)))
        scalar_names.append(k)

    print("\n" + "=" * 70)
    print("STEP 4: Train Ensemble Model (SVR + LGBM + XGB + ExtraTrees + Ridge)")
    
    from sklearn.svm import SVR
    from sklearn.ensemble import ExtraTreesRegressor, GradientBoostingRegressor, RandomForestRegressor
    from sklearn.neighbors import KNeighborsRegressor
    from sklearn.linear_model import BayesianRidge
    import xgboost as xgb
    from sklearn.linear_model import Ridge
    from sklearn.ensemble import ExtraTreesRegressor, AdaBoostRegressor
    from sklearn.preprocessing import PowerTransformer
    from catboost import CatBoostRegressor
    from sklearn.linear_model import HuberRegressor
    import lightgbm as lgb
    import xgboost as xgb
    from sklearn.base import clone

    from sklearn.preprocessing import MinMaxScaler

    base_models = [
        ('CatBoost_1', CatBoostRegressor(iterations=1000, learning_rate=0.03, depth=6, l2_leaf_reg=5.0, random_state=42, verbose=False)),
        ('XGB_1', xgb.XGBRegressor(n_estimators=500, max_depth=5, learning_rate=0.03, subsample=0.8, colsample_bytree=0.8, random_state=42)),
        ('LGBM_1', lgb.LGBMRegressor(n_estimators=500, learning_rate=0.03, num_leaves=15, subsample=0.8, colsample_bytree=0.8, random_state=42, verbose=-1)),
        ('SVR_RBF', SVR(kernel='rbf', C=5.0, gamma='scale', epsilon=0.1)),
        ('SVR_Poly', SVR(kernel='poly', C=2.0)),
        ('Ridge_1', Ridge(alpha=100.0)),
        ('BayesianRidge', BayesianRidge()),
        ('KNN_5', KNeighborsRegressor(n_neighbors=5)),
        ('KNN_15', KNeighborsRegressor(n_neighbors=15)),
        ('RF_1', RandomForestRegressor(n_estimators=500, max_depth=10, random_state=42)),
('ET_1', ExtraTreesRegressor(n_estimators=500, max_depth=10, random_state=42)),
        ('Huber', HuberRegressor())
    ]

    model_names = [n for n, _ in base_models] + ['PairwiseRankNN']
    oof_matrix = np.zeros((len(y_train), len(model_names)))

    for train_idx, val_idx in folds:
        X_tr_raw, X_va_raw = X_scalars[train_idx], X_scalars[val_idx]
        y_tr = y_train[train_idx]

        fold_scaler = MinMaxScaler()
        X_tr = fold_scaler.fit_transform(X_tr_raw)
        X_va = fold_scaler.transform(X_va_raw)

        for m_idx, (name, m) in enumerate(base_models):
            mc = clone(m)
            mc.fit(X_tr, y_tr)
            preds = mc.predict(X_va)
            oof_matrix[val_idx, m_idx] = preds

        rank_model_fold = train_pairwise_rank_model(
            X_tr,
            y_tr,
            X_val=X_va,
            y_val=y_train[val_idx],
            epochs=380,
            lambda_rank=0.80,
            lambda_mse=0.25,
            patience=35,
        )
        rank_preds = predict_pairwise_rank_model(rank_model_fold, X_va)
        oof_matrix[val_idx, len(base_models)] = rank_preds

    # Bounded Optimizing Weights (Prevents Test Set Drift while maximizing capability)
    from scipy.optimize import minimize
    def ensemble_loss(weights):
        w = weights / (np.sum(weights) + 1e-12)
        preds = np.average(oof_matrix, axis=1, weights=w)
        s = spearmanr(preds, y_train)[0]
        p = pearsonr(preds, y_train)[0]
        # Focus more on ranking while keeping a small concentration penalty.
        l2_penalty = 0.15 * np.sum((w - 1.0/len(w))**2)
        return -(0.75 * s + 0.25 * p) + l2_penalty

    init_w = np.ones(len(model_names)) / len(model_names)
    
    # Let weak models drop to 0 while keeping any single model from dominating too much.
    bounds = [(0.0, 0.35) for _ in range(len(model_names))]
    
    res = minimize(ensemble_loss, init_w, bounds=bounds, method='SLSQP')
    best_weights = res.x / (np.sum(res.x) + 1e-12)

    print(f"  Using Bounded Evaluated Weights: {dict(zip(model_names, np.round(best_weights, 3)))}")
    
    oof_final = np.average(oof_matrix, axis=1, weights=best_weights)
    final_s = spearmanr(oof_final, y_train)[0]
    final_p = pearsonr(oof_final, y_train)[0]
    final_target = 0.6 * final_s + 0.4 * final_p
    print(f"  CV Ensemble TARGET = {final_target:.4f}")

    # Retrain on full data
    scaler_full = MinMaxScaler()
    X_scalars_scaled_full = scaler_full.fit_transform(X_scalars)

    final_models = {}
    for name, m in base_models:
        mc = clone(m)
        mc.fit(X_scalars_scaled_full, y_train)
        final_models[name] = mc

    final_rank_model = train_pairwise_rank_model(
        X_scalars_scaled_full,
        y_train,
        epochs=460,
        lambda_rank=0.80,
        lambda_mse=0.25,
        patience=60,
    )
        
    print("\n" + "=" * 70)
    print("STEP 5: Validate and Predict on TEST SET")
    
    # Strictly target the provided test set images
    target_files = sorted(os.listdir(target_dist_dir))
    val_sys_preds = []
    
    for i, fname in enumerate(target_files):
        d_path = os.path.join(target_dist_dir, fname)
        r_path = os.path.join(target_ref_dir, fname)
        
        # Standard Pass
        dense, scalars = extract_features_cached(f"test_real_{fname}", d_path, r_path, models, flip=False)
        feat_vec = [scalars[k] for k in scalar_names if not k.endswith('_ridge_pred')]
        for k in ['clipH', 'clipL', 'dino', 'dino_patch_sim', 'dino_336', 'dino_patch_sim_336']:
            v = fitted_meta[k].predict(dense[k].reshape(1, -1))[0]
            feat_vec.append(v)
            
        feat_vec = np.nan_to_num(feat_vec)

        feat_vec = scaler_full.transform([feat_vec])[0]
        m_preds_1 = [m.predict([feat_vec])[0] for _, m in final_models.items()]
        rank_pred = predict_pairwise_rank_model(final_rank_model, np.array([feat_vec]))[0]
        m_preds_1.append(rank_pred)
        pred = np.average(m_preds_1, weights=best_weights)
        
        val_sys_preds.append((fname, pred))
        
        if (i+1) % 20 == 0:
            print(f"  Predicted {i+1} / {len(target_files)}")
            
    print("\n" + "=" * 70)
    print("STEP 6: Export Submission")
    
    sub_dir = os.path.abspath(args.output_dir)
    os.makedirs(sub_dir, exist_ok=True)
    
    xlsx_path = os.path.join(sub_dir, 'prediction.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['picture_name', 'Score'])
    for name, pred in val_sys_preds:
        # Scale protection just to be completely safe
        clipped_score = float(np.clip(pred, 0, 5))
        ws.append([name, clipped_score])
    wb.save(xlsx_path)

    readme_path = os.path.join(sub_dir, 'readme.txt')
    readme_bytes = resolve_readme_bytes(args.readme_template_zip, extra_data_flag=1)
    with open(readme_path, 'wb') as f:
        f.write(readme_bytes)

    zip_path = os.path.abspath(args.output_zip)
    write_submission_zip(zip_path, xlsx_path, readme_bytes)

    validate_submission_artifacts(zip_path, expected_rows=len(target_files))

    manifest_path = os.path.join(sub_dir, 'reproducibility_manifest.json')
    write_repro_manifest(
        manifest_path=manifest_path,
        args=args,
        device=DEVICE,
        train_count=len(train_files),
        test_count=len(target_files),
        metrics={
            'srocc': final_s,
            'plcc': final_p,
            'target': final_target,
        },
        xlsx_path=xlsx_path,
        zip_path=zip_path,
        elapsed_sec=time.time() - started,
    )

    print(f"Successfully wrote test predictions to: {zip_path}")
    print(f"Manifest written to: {manifest_path}")
    print("ALL DONE.")

if __name__ == '__main__':
    cli_args = build_arg_parser().parse_args()
    main(cli_args)